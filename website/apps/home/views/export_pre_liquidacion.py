from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook

class ExportPreLiquidacion(View):

    def post(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        data = self.get_data()
        ws['A1'] = 'Número de carga'
        ws['B1'] = 'Tienda'
        ws['C1'] = 'Número de bultos'
        ws['D1'] = 'Peso (Kg.)'
        ws['E1'] = 'A pagar cumple (S/.)'
        ws['F1'] = 'A pagar general (S/.)'
        i = 0
        for row in data:
            i += 1
            ws.append(row)
        ws = self.set_resume(ws, i)
        nombre_archivo = "reporte_pre_liquidacion.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

    def get_data(self):
        numero_carga = self.request.POST.getlist('numero_carga[]')
        tienda = self.request.POST.getlist('tienda[]')
        numero_bultos = self.request.POST.getlist('numero_bultos[]')
        pesos = self.request.POST.getlist('pesos[]')
        total = self.request.POST.getlist('total[]')
        item_paid_general = self.request.POST.getlist('item_paid_general[]')
        data_convert = list(zip(numero_carga, tienda, numero_bultos, pesos, total, item_paid_general))
        return data_convert

    def set_resume(self, ws, i):
        ws.cell(column=5, row=i + 3, value="Total a pagar cumple")
        ws.cell(column=6, row=i + 3, value=self.request.POST.get('gran_total'))
        ws.cell(column=5, row=i + 4, value="Total a pagar general")
        ws.cell(column=6, row=i + 4,
                value=self.request.POST.get('total_item_paid_general'))
        return ws

