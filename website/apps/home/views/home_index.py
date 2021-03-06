from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.views import View
from django.http import HttpResponseForbidden, HttpResponseBadRequest
from django.urls import reverse
from django.views.generic import FormView
from django.views.generic.detail import SingleObjectMixin
from django.shortcuts import render, render_to_response, redirect, get_object_or_404
from django.db.models import Max
import time
from datetime import datetime
from pytz import timezone
from django.views.generic import TemplateView
import json

from apps.home.forms.form import *

from django import forms
from django.template import RequestContext

from planificacion.models import *
from planificacion.punto_geo import PuntoGeo


import openpyxl
from django.db import IntegrityError
from django.core.exceptions import ValidationError

from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt

from django.db.models import Q
from django.contrib.auth.models import User,Group,Permission
from django.http import JsonResponse
from django.db.models import Max


#from django.shortcuts import get_object_or_404, redirect, render, reverse

from utils.utils import (
    get_max_ruta_codigo_planificacion_punto_control,
    get_list_ruta_codigo,
    get_list_lpn,
    get_match,
    get_inside,
    get_outside,
    get_dict_carga_bulto,
    get_planificacion_geocerca,
    get_num_planificacion_punto_control,
    get_carga_bulto_filtered,
    get_count_marcaciones,
    get_marcaciones_for_report,
    get_list_marcaciones_match,
)

class index(SingleObjectMixin, FormView):
    form_class = MyForm
    initial = {'key': 'value'}
    template_name = 'index.html'
    
    def get(self, request, *args, **kwargs):
        
        form = self.form_class(initial=self.initial)
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            # <process form cleaned data>
            return HttpResponseRedirect('/success/')

        return render(request, self.template_name, {'form': form})

def guardar_carga(request):
    import json
    zona = 'America/Lima'
    info = {}
    fecha = datetime.now(timezone(zona))
    bultos = PrePlanificacionCargaBulto.objects.filter(user=request.user)
    for b in bultos:
        PlanificacionCargaBulto.objects.create(user=b.user,fecha_carga=b.fecha_carga,numero_carga=b.numero_carga,
                        ruta_codigo=b.ruta_codigo,destino=b.destino,local=b.local,paradas=b.paradas,numero_lpn=b.numero_lpn,
                        peso=b.peso,volumen=b.volumen,guia_remision=b.guia_remision,transportista=b.transportista,
                        fecha_llegada=b.fecha_llegada,region=b.region,fecha_registro=b.fecha_registro,fecha_llegada_inicial=b.fecha_llegada_inicial,unidades=b.unidades)

    PrePlanificacionCargaBulto.objects.filter(user=request.user).delete()

    mCargaBulto = PlanificacionCargaBulto.objects.all()
    paginator = Paginator(mCargaBulto, 20) # Show 25 contacts per page

    page = request.GET.get('page')
    data = paginator.get_page(page)

    marcadores = PlanificacionCargaPuntoControl.objects.all()
    PuntoGeo.crearPuntoGeo(bultos,request.user.id,marcadores)
        
    return render(request, 'index.html', {"data1":data})

# Create your views here.
def import_data(request):
    message = ''
    status = None
    mCargaBulto = []
    if PuntoGeo.permisos(request.user.id).carga_masiva:
        
        if "GET" == request.method:
            status = None
            #return render(request, 'index.html', {})
        if "POST" == request.method:
            user = User.objects.get(id=request.user.id)
            

            try:
                excel_file = request.FILES["excel_file"]
                PrePlanificacionCargaBulto.objects.filter(user=request.user).delete()

                # you may put validations here to check extension or file size

                wb = openpyxl.load_workbook(excel_file)
                #worksheet = wb.get_sheet_by_name('Hoja1')
                worksheet = wb.active

                # getting a particular sheet by name out of many sheets

                excel_data = list()
                data = list()
                # iterating over the rows and
                # getting value from each cell in row
                import datetime
                import time
                import re
            
                i = 2
                for row in worksheet.iter_rows():

                    row_data = list()

                    fecha_carga = worksheet['A'+str(i)].value
                    if fecha_carga != None:
                        fecha_carga = worksheet['A'+str(i)].value
                        fecha_carga = datetime.datetime.strptime(str(fecha_carga), '%Y-%m-%d %H:%M:%S')

                        numero_carga = worksheet['B'+str(i)].value
                        ruta_codigo = worksheet['C'+str(i)].value
                        destino = worksheet['D'+str(i)].value
                        #local = worksheet['E'+str(i)].value
                        paradas = worksheet['E'+str(i)].value
                        numero_lpn = worksheet['F'+str(i)].value
                        peso = str(worksheet['G'+str(i)].value)
                        volumen = worksheet['H'+str(i)].value
                        guia_remision = worksheet['I'+str(i)].value
                        transportista = worksheet['J'+str(i)].value
                        fecha_llegada_inicial = worksheet['K'+str(i)].value
                        fecha_llegada = worksheet['L'+str(i)].value
                        if fecha_llegada != None:
                            fecha_llegada = datetime.datetime.strptime(str(fecha_llegada), '%Y-%m-%d %H:%M:%S')
                        if fecha_llegada_inicial != None:
                            fecha_llegada_inicial = datetime.datetime.strptime(str(fecha_llegada_inicial), '%Y-%m-%d %H:%M:%S')
                        region = worksheet['M'+str(i)].value
                        unidades = worksheet['N'+str(i)].value
                        now = datetime.datetime.now()
                        
                        if ruta_codigo != None:
                            try:
                                
                                objRuta = PlanificacionRuta.objects.get(codigo=ruta_codigo)
                            except:
                                objRuta = PlanificacionRuta(codigo=ruta_codigo)
                        else:
                            objRuta = PlanificacionRuta(codigo='')

                        mCargaBulto.append(PlanificacionCargaBulto(user=user,fecha_carga=fecha_carga,numero_carga=numero_carga,
                            ruta_codigo=objRuta,destino=destino,paradas=paradas,numero_lpn=numero_lpn,
                            peso=peso,volumen=volumen,guia_remision=guia_remision,transportista=transportista,
                            fecha_llegada=fecha_llegada,region=region,fecha_registro=now,fecha_llegada_inicial=fecha_llegada_inicial,unidades=unidades))
                            
                        i=i+1

                status = True
                message = 'El archivo se cargo correctamente.'
        

            except IntegrityError as e: 
                status = False
                message = 'Lo sentimos, problemas al cargar el archivo. ' + str(e)
            except ValidationError as e:
                status = False
                message = 'Lo sentimos, problemas al cargar el archivo. ' + str(e)
            except ValueError as e:
                status = False
                message = 'Lo sentimos, problemas al cargar el archivo. ' + str(e)
            except Exception as e:
                status = False
                message = 'Lo sentimos, problemas al cargar el archivo. ' + str(e)

            

    #**************Validacionees****************************************************

            mCargaBultoCopy = mCargaBulto.copy()

            cont = 0
            lpn = []
            lpnbd = []
            lpnempty = []
            strlpnempty = ''
            strlpn = ''
            strlpnbd = ''
            regcont = 1
            geo = []
            strgeo = ''
            ruta = []
            strruta = ''
            
            for mb in mCargaBulto:
                regcont = regcont + 1
                rutas = 10 
                geos = 10
                lpns = PlanificacionCargaBulto.objects.filter(numero_lpn=mb.numero_lpn).count()
                if mb.destino != None:
                    geos = PlanificacionGeocerca.objects.filter(codigo=mb.destino).count()

                if mb.ruta_codigo.codigo != '':
                    rutas = PlanificacionPuntoControl.objects.filter(nombre=mb.ruta_codigo).count()
                    
                if rutas == 0:
                    ruta.append(mb.ruta_codigo)

                if geos == 0:
                    geo.append(mb.destino)
                
                if (mb.fecha_carga == None) or (mb.numero_carga == None) or (mb.ruta_codigo.codigo == '') or (mb.destino == None):
                    lpnempty.append(regcont)

                if lpns > 0:
                    lpnbd.append(mb.numero_lpn)
                for mbc in mCargaBultoCopy:
                    if mb.numero_lpn == mbc.numero_lpn:
                        cont = cont + 1
                        if cont >= 2:
                            lpn.append(mb.numero_lpn)
                cont = 0

            message1 = ''
            message2 = ''
            message3 = ''
            message4 = ''
            message5 = ''

            if len(lpn) > 0:
                result = list(set(lpn))
                for r in result:
                    strlpn = strlpn+str(r)+','
                
                status = False
                message1 = 'Los siguientes campos (numero_lpn) estas duplicados: '+strlpn

            if len(lpnbd) > 0:
                result = list(set(lpnbd))
                for r in result:
                    strlpnbd = strlpnbd+str(r)+','
                
                status = False 
                message2 = 'Los siguientes campos (numero_lpn) ya existen en la base de datos: '+strlpnbd

            if len(lpnempty) > 0:
                for r in lpnempty:
                    strlpnempty  = strlpnempty +str(r)+','   
                
                status = False
                message3 = 'Los siguientes registros tienen campos obligatorios vacíos: '+strlpnempty 

            if len(geo) > 0:
                result = list(set(geo))
                for r in result:
                    strgeo  = strgeo +str(r)+',' 
                
                status = False
                message4 = 'Los siguientes campos (destino) no existen como código de Geocerca: '+strgeo

            if len(ruta) > 0:
                result = list(set(ruta))
                for r in result:
                    strruta  = strruta +str(r)+','
                
                status = False
                message5 = 'Los siguientes campos (ruta) no están asociados a un punto de control: '+strruta 

            if not status:
                message = 'Lo sentimos, no es posible guardar este archivo. '+message1+". "+message2+". "+message3+". "+message4+". "+message5
            else:
                for b in mCargaBulto:
                    PrePlanificacionCargaBulto.objects.create(user=b.user,fecha_carga=b.fecha_carga,numero_carga=b.numero_carga,
                            ruta_codigo=b.ruta_codigo,destino=b.destino,paradas=b.paradas,numero_lpn=b.numero_lpn,
                            peso=b.peso,volumen=b.volumen,guia_remision=b.guia_remision,transportista=b.transportista,
                            fecha_llegada=b.fecha_llegada,region=b.region,fecha_registro=b.fecha_registro,fecha_llegada_inicial=b.fecha_llegada_inicial,unidades=b.unidades)

            

    #***********************************************************************************    
        
        paginator = Paginator(mCargaBulto, 20) # Show 25 contacts per page

        page = request.GET.get('page')
        data = paginator.get_page(page)


        return render(request, 'index.html', {"data":data,"message":message,"status":status})

    else:
        return render(request, 'permisos.html')

@csrf_exempt
def eliminar_carga(request):
    import json
    message = ''
    success = True
    data = request.POST.getlist('data[]')

    try:
        i = 0
        while i < len(data):

            PlanificacionCargaBulto.objects.filter(id=data[i]).delete()
            #PlanificacionCargaBulto.objects.create(id=datas[i].get('id'))
            i = i + 1
        message = str(i) + ' Registro(s) eliminado(s).'
    except PlanificacionCargaBulto.DoesNotExist:
        message = 'Error al eliminar.'
        success = False
    except ValidationError as e: 
        success = False
        message = 'Lo sentimos, problemas al cargar el archivo. ' + str(e)

    return render(request, 'index.html', {"message":message,"status":success})

@csrf_exempt
def crear_geocercas(request):
    import json
    message = ''
    success = True
    form = planificacionGeocercaForm()
    if "POST" == request.method:

        codigo = request.POST.get('codigo')
        nombre = request.POST.get('nombre')
        direccion = request.POST.get('direccion')
        tipo_geocerca = request.POST.get('tipo_geocerca')
        coordenadas = request.POST.get('coordenadas')
        coordenadas_poligon = request.POST.get('coordenadas_poligon')

        if coordenadas:
            c = json.loads(coordenadas)
            try:
                PlanificacionGeocerca.objects.get(codigo=codigo)
                success = False
                message = 'Error, el código de geocerca ya existe'
            except PlanificacionGeocerca.DoesNotExist:
                try:
                    form = planificacionGeocercaForm(request.POST)
                    message = form.errors
                    if form.is_valid():
                        #form.save()
                        mTG = PlanificacionTipoGeocerca.objects.get(id=tipo_geocerca)
                        mPG = PlanificacionGeocerca.objects.create(codigo=codigo,
                            nombre=nombre,direccion=direccion,
                            tipo_geocerca=mTG,
                            coordenadas=coordenadas,
                            coordenadas_poligon=coordenadas_poligon)
                        i = 0
                        while i < len(c):
                            PlanificacionDetalleGeocerca.objects.create(geocerca_id=mPG.id,latitud=c[i]['lat'],longitud=c[i]['lng'])
                            i = i + 1
                        message = 'La Geocerca se creo correctamente.'
                        return redirect('/listar_geocercas/')
                except PlanificacionGeocerca.DoesNotExist:
                    message = 'Error tabla no existe.'
                    success = False
                except ValidationError as e:
                    success = False
                    message = 'Lo sentimos, problemas al registrar.' + str(e)
                except IntegrityError as e:
                    success = False
                    message = 'Lo sentimos, problemas al registrar. ' + str(e)
    return render(request, 'geocercas.html', {"message":message,"status":success,"form":form})

@csrf_exempt
def listar_geocercas(request):
    if PuntoGeo.permisos(request.user.id).geocercas:
        data = None
        message = ''
        status = None
        if "POST" == request.method:

            nom_geo = request.POST.get('nom_geo')
            cod_geo = request.POST.get('cod_geo')
            query = {}

            if nom_geo != '':
                query['nombre'] = nom_geo

            if cod_geo != '':
                query['codigo'] = cod_geo

            if nom_geo != '' or cod_geo != '':
                dataGeocerca = PlanificacionGeocerca.objects.filter(**query)
            else:
                dataGeocerca = PlanificacionGeocerca.objects.all()

            # paginator = Paginator(dataGeocerca, 20) # Show 25 contacts per page

            # page = request.GET.get('page')
            # data = paginator.get_page(page)

            data = dataGeocerca

        return render(request, 'listar_geocercas.html',
                      {"data": data, "message": message, "status": status})
    else:
        return render(request, 'permisos.html')


@csrf_exempt
def eliminar_geocercas(request):
    import json
    message = ''
    status = False
    data = request.POST.getlist('data[]')

    try:
        i = 0
        while i < len(data):
            PlanificacionGeocerca.objects.filter(id=data[i]).delete()
            #PlanificacionCargaBulto.objects.create(id=datas[i].get('id'))
            i = i + 1
        message = str(i) + ' Registro(s) eliminado(s).'
        status = True
    except PlanificacionGeocerca.DoesNotExist:
        message = 'Lo sentimos, problemas al eliminar.'
        status = False
    except IntegrityError as e: 
        status = False
        message = 'Lo sentimos, problemas al eliminar. ' + str(e)

    dataGeocerca = PlanificacionGeocerca.objects.all()

    paginator = Paginator(dataGeocerca, 20) # Show 25 contacts per page

    page = request.GET.get('page')
    data = paginator.get_page(page)

    return render(request, 'listar_geocercas.html', {"data":data,"message":message,"status":status})

    #return redirect('/listar_geocercas/')
    #return render(request, 'listar_geocercas.html', {"message":message,"status":success})


# @csrf_exempt
# def edita_geocercas(request, id=None, template_name='editar_geocercas.html'):
#     if id:
#         geocerca = get_object_or_404(PlanificacionGeocerca, pk=id)
#
#         if not geocerca.id:
#             return HttpResponseForbidden()
#     #else:
#     #    article = Article(author=request.user)
#
#     form = planificacionGeocercaForm(request.POST or None, instance=geocerca)
#     c = PlanificacionDetalleGeocerca.objects.filter(geocerca_id=id)
#
#     coordenadas = ''
#     coordenadas_poligon = ''
#     i = 0
#     while i < len(c):
#         coordenadas += '{ "lat": '+c[i].latitud+', "lng": '+c[i].longitud+' }, '
#         coordenadas_poligon += c[i].latitud+' '+c[i].longitud+', '
#         i = i + 1
#
#     cp = coordenadas_poligon[0 : -2]
#     coordenadas_poligon = '('+geocerca.coordenadas_poligon+')'
#
#     if "POST" == request.method:
#         if request.POST and form.is_valid():
#             form.save()
#
#             # Save was successful, so redirect to another page
#             redirect_url = reverse(geocerca_save_success)
#             return redirect(redirect_url)
#
#     return render(request, template_name, {
#         'form': form, 'id':id,'coordenadas_poligon':coordenadas_poligon
#     })

@csrf_exempt
def editgeo(request, id):
    import json

    geocerca = PlanificacionGeocerca.objects.get(id=id)
    success = True
    message = ''

    if "POST" == request.method:

        coordenadas = request.POST.get('coordenadas')
        coordenadas_poligon = request.POST.get('coordenadas_poligon')
        codigo = request.POST.get('codigo')

        form = planificacionGeocercaForm(request.POST or None,
                                         instance=geocerca)
        cods = PlanificacionGeocerca.objects.filter(codigo=codigo).exclude(
            id=id)
        if len(cods) > 0:
            success = False
            message = 'Error, lo sentimos el código de geocerca ya existe'
        else:
            if request.POST and form.is_valid():
                form.save()

                mG = PlanificacionGeocerca.objects.get(id=id)
                mG.coordenadas_poligon = coordenadas_poligon
                mG.save()

                c = json.loads(coordenadas)
                try:
                    mPG = PlanificacionDetalleGeocerca.objects.filter(
                        geocerca_id=id)
                    m = 0
                    while m < len(mPG):
                        PlanificacionDetalleGeocerca.objects.get(
                            id=mPG[m].id).delete()
                        m = m + 1
                    i = 0
                    while i < len(c):
                        PlanificacionDetalleGeocerca.objects.create(
                            geocerca_id=geocerca.id, latitud=c[i]['lat'],
                            longitud=c[i]['lng'])
                        i = i + 1

                    message = 'La Geocerca se creo correctamente.'
                    return redirect('/listar_geocercas/')
                except PlanificacionGeocerca.DoesNotExist:
                    message = 'Error tabla no existe.'
                    success = False
                except ValidationError as e:
                    success = False
                    message = 'Lo sentimos, problemas al registrar.' + str(e)
                except IntegrityError as e:
                    success = False
                    message = 'Lo sentimos, problemas al registrar. ' + str(e)

                    # Save was successful, so redirect to another page
                    # return redirect('/listar_geocercas/', product_id=True)
    else:
        form = planificacionGeocercaForm(instance=geocerca)
    return render(request, 'geocercas_edit.html',
                  {"message": message, "status": success, "form": form})

def puntos_control(request):
    message = ''
    status = None
    dataRuta = PlanificacionRuta.objects.all()

    paginator = Paginator(dataRuta, 20) # Show 25 contacts per page

    page = request.GET.get('page')
    data = paginator.get_page(page)

    return render(request, 'puntos_control.html', {"data":data,"message":message,"status":status})

def point_in_poly(x,y,poly):
    """Return True if (x, y) lies within poly and False otherwise.
 
poly := [ (x1, y1), (x2, y2), (x3, y3), ... (xn, yn) ]
This uses the ray-casting algorithm for determining whether a
point is inside a polygon.
http://geospatialpython.com/2011/01/point-in-polygon.html
http://en.wikipedia.org/wiki/Point_in_polygon#Ray_casting_algorithm """
 
 
    n = len(poly)
    inside = False
 
    p1x,p1y = poly[0]
    for i in range(n+1):
        p2x,p2y = poly[i % n]
        if y > min(p1y,p2y):
            if y <= max(p1y,p2y):
                if x <= max(p1x,p2x):
                    if p1y != p2y:
                        xints = (y-p1y)*(p2x-p1x)/(p2y-p1y)+p1x
                    if p1x == p2x or x <= xints:
                        inside = not inside
        p1x,p1y = p2x,p2y
 
    return inside


class Upmodal1(TemplateView):
    def get(self, request,lpn,cnt):
        det = []
        if cnt == 'fuera':
            estado = 'Fuera'
            marcaciones = []
            
            todas = MarcacionesMatch.objects.filter(lpn=lpn).filter(dentro=0)
            matchs = []
            for t in todas:
                todas1 = MarcacionesMatch.objects.filter(lpn=lpn).filter(dentro=1).filter(id_marcacion = t.id_marcacion)[0:100]
                if len(todas1) == 0:
                    if not t.id_marcacion in marcaciones:
                        marcaciones.append(t.id_marcacion)
                        matchs.append(t)

            for mt in matchs:
                geo = PuntoGeo.hallarGeo(mt.punto)
                dic = {}
                dic['lpn'] = mt.lpn
                dic['punto'] = mt.punto
                dic['fecha'] = mt.fecha_marca
                if geo != None:
                    dic['nombre'] = geo.nombre
                    dic['dir'] = geo.direccion
                else:
                    dic['nombre'] = '-'
                    dic['dir'] = 'http://maps.google.com/?q='+mt.punto

                det.append(dic)

        else:
            cnt = cnt.replace('x','')
            estado = 'Dentro'
            if cnt == 'f':
                args = MarcacionesMatch.objects.filter(lpn=lpn)
                max = args.aggregate(Max('id_control'))
                matchs = MarcacionesMatch.objects.filter(lpn=lpn).filter(dentro=1).filter(id_control=max['id_control__max'])
            else:
                matchs = MarcacionesMatch.objects.filter(lpn=lpn).filter(dentro=1).filter(id_control=int(cnt))
        
            
            for mt in matchs:
                dic = {}
                if cnt == 'f':
                    dest = PlanificacionCargaBulto.objects.get(numero_lpn=mt.lpn)
                    geo = PlanificacionGeocerca.objects.get(codigo = dest.destino)
                    dic['nombre'] = geo.nombre
                else:
                    geo = PlanificacionPuntoControl.objects.get(ruta_codigo = mt.cnt_nombre,orden = mt.id_control)
                    dic['nombre'] = geo.geocerca.nombre

                dic['control'] = mt.id_control
                
                dic['punto'] = mt.punto
                dic['fecha'] = mt.fecha_marca
                det.append(dic)

        equis = {}
        equis['equis'] = det
        equis['estado'] = estado  
        return render(request, 'modal1.html',{'equis':equis})

@csrf_exempt
def Excel(request):
    data1 = request.POST.getlist('data1[]')
    data2 = request.POST.getlist('data2[]')
    data3 = request.POST.getlist('data3[]')
    data4 = request.POST.getlist('data4[]')
    data5 = request.POST.getlist('data5[]')
    data6 = request.POST.getlist('data6[]')
    data7 = request.POST.getlist('data7[]')
    data8 = request.POST.getlist('data8[]')
    control = request.POST.getlist('control[]')
    om = request.POST.getlist('om[]')
    # Export excel
    from openpyxl import Workbook
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.styles import colors
    from openpyxl.cell import Cell
    from datetime import datetime
    from pytz import timezone

    fmt = "%d/%m/%Y %H:%M"

    wb = Workbook()
    ws = wb.active

    redFill = PatternFill(start_color='FFF61723', end_color='FFF61723',
                          fill_type='solid')
    greenFill = PatternFill(start_color='FF07A34F', end_color='FF07A34F',
                            fill_type='solid')
    yellowFill = PatternFill(start_color='FFF8E903', end_color='FFF8E903',
                             fill_type='solid')
    greyFill = PatternFill(start_color='FFb3b3b3', end_color='FFb3b3b3',
                           fill_type='solid')

    # --------------DETALLE................
    ws['A1'] = 'Número de carga'
    ws['B1'] = 'Número de bandeja'
    ws['C1'] = '#Controles'
    ws['D1'] = 'Paradas'
    ws['E1'] = 'Ruta'
    ws['F1'] = 'Destino'
    ws['G1'] = 'Fecha de carga'
    ws['H1'] = 'Transportista'
    cnt = []
    equis = []
    for cn in control:
        arreglo = cn.replace('[', '').replace(']', '').split(',')
        cnt.append(arreglo)

    # for c1 in cnt[0]:
    #     equis.append(0)

    # for c2 in cnt:
    #     i = 0
    #     for c in c2:
    #         if 'x' in c and equis[i] == 0:
    #             equis[i] = 1
    #         i = i + 1

    c = 9
    i = 0
    for ct in cnt[0]:
        if i < len(ct):
            if i == 0:
                ws.cell(row=1, column=c).value = 'CD'

                ws.cell(row=1, column=c + 1).value = 'CD'
                c = c + 1
            else:
                ws.cell(row=1, column=c).value = 'Control ' + str(i)

                ws.cell(row=1, column=c + 1).value = 'Control ' + str(i)
                c = c + 1

        else:
            ws.cell(row=1, column=c).value = 'Local'
            ws.cell(row=1, column=c + 1).value = 'Local'
            c = c + 1
        i = i + 1
        c = c + 1

    longitudOM = 0
    for o in om:
        if int(o) > longitudOM:
            longitudOM = int(o)
    j = 0
    if longitudOM == 0:
        ws.cell(row=1, column=c + j).value = 'Otras Marcas'
    else:
        for i in range(longitudOM):
            ws.cell(row=1, column=c + j).value = 'Otras Marcas ' + str(i + 1)
            ws.cell(row=1, column=c + j + 1).value = 'Otras Marcas ' + str(
                i + 1)
            j = j + 2

    cont = 2

    for row1, row2, row3, row4, row5, row6, row7, row8, cn, o in zip(data1, data2,
                                                               data3, data4,
                                                               data5, data6,
                                                               data7, data8, cnt, om):
        if row2 != '':
            ws.cell(row=cont, column=1).value = row1
            ws.cell(row=cont, column=2).value = row2
            ws.cell(row=cont, column=3).value = row3
            ws.cell(row=cont, column=4).value = row4
            ws.cell(row=cont, column=5).value = row5
            ws.cell(row=cont, column=6).value = row6
            ws.cell(row=cont, column=7).value = row7
            ws.cell(row=cont, column=8).value = row8
            clm = 9
            k = 0
            for c in cn:
                #k = k + 1
                if 'x' in c:
                    #marca = MarcacionesMatch.objects.filter(lpn=row2, dentro=1,
                    #                                        id_control=k).first()
                    marca = MarcacionesMatch.objects.filter(lpn=row2, dentro=1).first()
                    ws.cell(row=cont,
                            column=clm).value = marca.fecha_marca.strftime(fmt) if marca and marca.fecha_marca else ''
                    ws.cell(row=cont, column=clm + 1).value = marca.punto if marca else ''
                    ws.cell(row=cont, column=clm).fill = greenFill
                    ws.cell(row=cont, column=clm + 1).fill = greenFill
                    clm = clm + 1
                elif '.' in c:
                    ws.cell(row=cont, column=clm).value = ''
                    ws.cell(row=cont, column=clm + 1).value = ''
                    ws.cell(row=cont, column=clm).fill = greyFill
                    ws.cell(row=cont, column=clm + 1).fill = greyFill
                    clm = clm + 1
                elif '-' in c:
                    ws.cell(row=cont, column=clm).value = ''
                    ws.cell(row=cont, column=clm + 1).value = ''
                    ws.cell(row=cont, column=clm).fill = redFill
                    ws.cell(row=cont, column=clm + 1).fill = redFill
                    clm = clm + 1
                else:
                    ws.cell(row=cont, column=clm).value = ''
                    ws.cell(row=cont, column=clm + 1).value = ''
                    clm = clm + 1
                clm = clm + 1

            j = 0
            marc = []
            todas = MarcacionesMatch.objects.filter(lpn=row2).filter(dentro=0)
            for t in todas:
                todas1 = MarcacionesMatch.objects.filter(lpn=row2).filter(
                    dentro=1).filter(id_marcacion=t.id_marcacion)
                if len(todas1) == 0:
                    if not t.id_marcacion in marc:
                        marc.append(t.id_marcacion)

            for ma in marc:

                ska = PlanificacionCargaPuntoControl.objects.filter(id=ma).first()
                ws.cell(row=cont,
                        column=clm + j).value = ska.fecha_registro.strftime(fmt) if ska and ska.fecha_registro else ''
                ws.cell(row=cont, column=clm + j).fill = yellowFill
                ws.cell(row=cont,
                        column=clm + j + 1).value = ska.latitud + ',' + ska.longitud if ska else ''
                ws.cell(row=cont, column=clm + j + 1).fill = yellowFill
                j = j + 2
            cont = cont + 1

    nombre_archivo = "Detalles.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

@csrf_exempt
def pre_liquidacion(request):
    if PuntoGeo.permisos(request.user.id).pre_liquidacion:
        import decimal
        detalles = []
        gran_total = decimal.Decimal('0.00')
        total_item_paid_general = decimal.Decimal('0.00')
        total = decimal.Decimal('0.00')
        val_zero = decimal.Decimal('0.00')
        if "POST" == request.method:
            numero_carga = request.POST.get('numero_carga')
            fecini = request.POST.get('fecini')
            fecfin = request.POST.get('fecfin')

            fecini = datetime.strftime(datetime.strptime(fecini,'%d/%m/%Y %H:%M:%S'),'%Y-%m-%d %H:%M:%S')
            fecfin = datetime.strftime(datetime.strptime(fecfin,'%d/%m/%Y %H:%M:%S'),'%Y-%m-%d %H:%M:%S')

            query = {}
            query['fecha_carga__range'] = [fecini, fecfin]

            if numero_carga != '':
                query['numero_carga'] = numero_carga

            if not request.user.is_staff:
                query['transportista'] = request.user.username
            
            carga_bulto = PlanificacionCargaBulto.objects.filter(**query)

            nuevo = []
            for cb in carga_bulto:
                if not cb.numero_carga+"-"+str(cb.destino) in nuevo:
                    nuevo.append(cb.numero_carga+"-"+str(cb.destino))

            
            for new in nuevo:
                dic = {}
                val = new.split("-")
                
                dic['numero_carga'] = val[0]
                dic['tienda'] = val[1]
                rows = PlanificacionCargaBulto.objects.filter(numero_carga=val[0]).filter(destino=val[1])
                n_bultos = len(rows)
                dic['numero_bultos'] = n_bultos
                peso = 0
                cont = 0
                for row in rows:
                    peso = peso + row.peso
                    marca_final = len(PlanificacionPuntoControl.objects.filter(ruta_codigo = row.ruta_codigo))
                    marcaciones = len(MarcacionesMatch.objects.filter(lpn=row.numero_lpn).filter(id_control=marca_final).filter(dentro=1))
                    if marcaciones > 0:
                        cont = cont + 1

                tarifa = PlanificacionGeocerca.objects.get(codigo=val[1]).tarifa or val_zero
                if cont == n_bultos:
                    total = round(decimal.Decimal(peso) * tarifa, 2)
                else:
                    item_paid_general = round(decimal.Decimal(peso) * tarifa, 2)

                dic['pesos'] = round(peso,2)
                dic['tarifa'] = tarifa
                dic['total'] = total
                dic['item_paid_general'] = item_paid_general
                gran_total = gran_total + total
                total_item_paid_general = total_item_paid_general + item_paid_general
                detalles.append(dic)
            




        return render(request, 'pre_liquidacion.html',
                      {'detalles':detalles,
                       'gran_total':gran_total,
                       'total_item_paid_general': total_item_paid_general})
    else:
        return render(request, 'permisos.html')

@csrf_exempt
def consulta_registros(request):
    if PuntoGeo.permisos(request.user.id).consulta_registros:
        from django.db.models import Prefetch
        numero_carga = request.POST.get('numero_carga')
        numero_bandeja = request.POST.get('numero_bandeja')
        ruta = request.POST.get('ruta')
        destino = request.POST.get('destino')
        trans = request.POST.get('trans')
        fecini = request.POST.get('fecini')
        fecfin = request.POST.get('fecfin')
        resumen = {}
        detalles = []

        if "POST" == request.method:

            fecini = datetime.strftime(datetime.strptime(fecini,'%d/%m/%Y %H:%M:%S'),'%Y-%m-%d %H:%M:%S')
            fecfin = datetime.strftime(datetime.strptime(fecfin,'%d/%m/%Y %H:%M:%S'),'%Y-%m-%d %H:%M:%S')

            query = {}
            query['fecha_carga__range'] = [fecini, fecfin]
            if numero_carga != '':
                query['numero_carga'] = numero_carga
            if numero_bandeja != '':
                query['numero_lpn'] = numero_bandeja
            if ruta != '':
                query['ruta_codigo'] = ruta
            if destino != '':
                query['destino'] = destino
            if trans != '':
                query['transportista'] = trans

            if not request.user.is_staff:
                query['transportista'] = request.user.username

            #import pdb; pdb.set_trace()
            carga_bulto = PlanificacionCargaBulto.objects.filter(**query).order_by('numero_carga','destino')
            #max_puntos = 0
            # for cb in carga_bulto:
            #     print (cb)
            #     plan_ruta = len(PlanificacionPuntoControl.objects.filter(ruta_codigo = cb.ruta_codigo))
            #     if max_puntos < plan_ruta:
            #         max_puntos = plan_ruta
            lista_ruta_codigo = get_list_ruta_codigo(carga_bulto)
            max_puntos = get_max_ruta_codigo_planificacion_punto_control(lista_ruta_codigo)
            #print ('paso!!', max_puntos)

            control = []
            total = []
            dentro = []
            fuera = []
            #control.append(0)
            control.append('Objetivo')
            #total.append(len(carga_bulto))
            total.append(carga_bulto.count())
            dentro.append('')
            fuera.append('')

            n_carga = []
            n_lpn = []
            n_controles = []
            ruta = []
            f_carga = []
            cnt =[]
            trans = []

            dic = {}

            contf = 0
            contf1 = 0
            contf2 = 0
            control.append('CD')
            control_ = ["Local" if max_puntos==x else 'Control ' + str(x - 1) for x in range(2, max_puntos + 1)]
            control += control_
            list_id_control = [x for x in range(1, max_puntos + 1)]
            # for i in range(1,max_puntos + 1):
            #     if i < max_puntos:
            #         control.append(i)
            #     else:
            #         control.append('final')
            #     cont = 0
            #     cont1 = 0
            #     cont2 = 0
                #for cb in carga_bulto:

                    #orden = len(PlanificacionPuntotk_v2_2.2.2.1.1Control.objects.filter(ruta_codigo = cb.ruta_codigo))

            # list_marcaciones_match = list(
            #     MarcacionesMatch.objects.filter(
            #         fecha_marca__range=[fecini, fecfin]).values().all())

            l = get_list_lpn(carga_bulto)
            list_marcaciones_match = get_list_marcaciones_match(l)
            #match = get_match(l, list_id_control)
            match = [get_match(list_marcaciones_match, l, idc) for idc in list_id_control]
            #inside = get_inside(list_marcaciones_match, l, list_id_control)
            inside = [get_inside(list_marcaciones_match, l, idc) for idc in list_id_control]
            #outside = get_outside(list_marcaciones_match, l, list_id_control)
            outside = [get_outside(list_marcaciones_match, l, idc) for idc in list_id_control]
            total += match
            dentro += inside
            fuera += outside
            numCarga = []

            carga_bulto = get_dict_carga_bulto(carga_bulto)
            list_planificacion_geocerca = list(PlanificacionGeocerca.objects.values().all())
            list_planificacion_punto_control = list(PlanificacionPuntoControl.objects.values().all())

            for cb in carga_bulto:
                dic = {}
                innercnt = []
                geo = get_planificacion_geocerca(list_planificacion_geocerca, cb.get('destino'))
                nc = get_num_planificacion_punto_control(
                    list_planificacion_punto_control, cb.get('ruta_codigo'))
                if not cb.get('numero_carga') + "-" + str(cb.get('destino')) in numCarga:
                    dic1 = {}
                    innercnt1 = []
                    numCarga.append(cb.get('numero_carga') + "-" + str(cb.get('destino')))
                    dic1['numero_carga'] = cb.get('numero_carga') + "-" + str(
                        cb.get('destino'))
                    dic1['numero_lpn'] = ''
                    dic1['numero_controles'] = nc
                    dic1['paradas'] = cb.get('paradas')
                    dic1['ruta_codigo'] = cb.get('ruta_codigo')
                    #dic1['destino'] = geo.nombre
                    dic1['destino'] = geo.get('nombre', '')
                    dic1['fecha_carga'] = cb.get('fecha_carga')
                    dic1['transportista'] = cb.get('transportista')
                    dic1['om'] = 0
                    dic1['display'] = 'table-row'
                    orden = get_num_planificacion_punto_control(list_planificacion_punto_control, cb.get('ruta_codigo'))
                    bultos = get_carga_bulto_filtered(carga_bulto, cb.get('numero_carga'), cb.get('destino'))
                    all = 0
                    for i in range(1,max_puntos + 1):
                        con_in = 0
                        for b in bultos:
                            match = get_count_marcaciones(list_marcaciones_match, b.get('numero_lpn'), i, 1)
                            match1 = get_count_marcaciones(list_marcaciones_match, b.get('numero_lpn'), i, 0)
                            if len(match) > 0 and len(match1) >= 0:
                                con_in = con_in + 1
                        if i < orden :
                            if len(bultos) == con_in:
                                innercnt1.append('all')
                            else:
                                innercnt1.append('')

                        elif i >= orden and i < max_puntos:
                            if len(bultos) == con_in:
                                innercnt1.append('')
                                all = 1
                            else:
                                innercnt1.append('')

                        elif i == orden and i == max_puntos:
                            if len(bultos) == con_in:
                                all = 1

                        if i == max_puntos:
                            if all == 1:
                                innercnt1.append('all')
                            else:
                                innercnt1.append('')

                    dic1['control'] = innercnt1
                    detalles.append(dic1)


                #nc = get_num_planificacion_punto_control(list_planificacion_punto_control, cb.get('ruta_codigo'))
                dic['numero_carga'] = cb.get('numero_carga') + "-" + str(cb.get('destino'))
                dic['numero_lpn'] = cb.get('numero_lpn')
                dic['numero_controles'] = nc
                dic['ruta_codigo'] = cb.get('ruta_codigo')
                dic['destino'] = geo.get('nombre')
                dic['fecha_carga'] = cb.get('fecha_carga')
                dic['transportista'] = cb.get('transportista')
                dic['paradas'] = cb.get('paradas')
                orden = get_num_planificacion_punto_control(list_planificacion_punto_control, cb.get('ruta_codigo'))
                contx = 0
                cont_ = 0
                found = 0
            
                for i in range(1,max_puntos + 1):
                    match = get_count_marcaciones(list_marcaciones_match, cb.get('numero_lpn'), i, 1)
                    match1 = get_count_marcaciones(list_marcaciones_match, cb.get('numero_lpn'), i, 0)
                    if i < orden:
                        if len(match) > 0:
                            innercnt.append('x'+str(i))
                        elif len(match1) > 0:
                            innercnt.append('-')
                        else:
                            innercnt.append('')
                            found = 1
                    elif i == orden and i < max_puntos:
                        if len(match) > 0 :
                            innercnt.append('.')
                            contx = 1
                        elif len(match1) > 0:
                            innercnt.append('.')
                            cont_ = 1
                        else:
                            innercnt.append('')
                            found = 1
                    elif i == orden and i == max_puntos:
                        if len(match) > 0 :
                            contx = 1
                        elif len(match1) > 0:
                            cont_ = 1
                        else:
                            found = 1

                    elif i > orden and i < max_puntos:
                        innercnt.append('')

                    if i == max_puntos:
                        if contx > 0:
                            innercnt.append('xf')
                        elif cont_ > 0:
                            innercnt.append('-')
                        else:
                            innercnt.append('')

                marcaciones = []
                todas = get_count_marcaciones(list_marcaciones_match, cb.get('numero_lpn'), '', 0)
                cont = 0
                for t in todas:
                    todas1 = get_marcaciones_for_report(list_marcaciones_match, cb.get('numero_lpn'), t.get('id_marcacion'), 1)
                    if len(todas1) == 0:
                        if not t.get('id_marcacion') in marcaciones:
                            marcaciones.append(t.get('id_marcacion'))
                            cont = cont + 1

                dic['om'] = cont
                dic['control'] = innercnt

                dic['display'] = 'table-row'
                if not cb.get('numero_carga') + "-" + str(cb.get('destino')) in numCarga:
                    numCarga.append(cb.get('numero_carga') + "-" + str(cb.get('destino')))
                else:
                    dic['display'] = 'none'
                detalles.append(dic)

            resumen['detalles'] = detalles
            resumen['control'] = control
            resumen['total'] = total
            resumen['dentro'] = dentro
            resumen['fuera'] = fuera
    
        return render(request, 'consulta_registros.html', {'resumen':resumen})
    else:
        return render(request, 'permisos.html')

@csrf_exempt
def consulta_registros1(request):
    from django.db.models import Prefetch
    numero_carga = request.GET.get('numero_carga')
    numero_bandeja = request.GET.get('numero_bandeja')
    ruta = request.GET.get('ruta')
    local = request.GET.get('local')
    fecini = request.GET.get('fecini')
    fecfin = request.GET.get('fecfin')

    data_pc = {}
    countcargainicial = {}
    resum = []
    detalles = {}
    detallado = {}
    message = ''
    status = False
    data = []
    resumen = {}
    data2 = []
    control = []

    if "GET" == request.method:
        from django.db.models import Count
        #from shapely import wkt
        try:
            dataPuntoControl = PlanificacionPuntoControl.objects.all().select_related('geocerca').select_related('tipo_control').select_related('ruta_codigo').order_by('-orden')
            data_pc = PlanificacionPuntoControl.objects.values('ruta_codigo').annotate(total=Count('ruta_codigo')).order_by('-total').first()

            countcargainicial = PlanificacionCargaBulto.objects.all().count()
            '''resumen = PlanificacionPuntoControl.objects.raw("select distinct t1.numero_lpn,t3.id,t3.nombre, "
                "t3.id, t1.ruta_codigo, t2.coordenadas_poligon point, t4.coordenadas_poligon polygon "
                "from planificacion_carga_bulto t1 "
                "inner join planificacion_carga_punto_control t2 on t1.numero_lpn = t2.numero_lpn "
                "inner join planificacion_punto_control t3 on t1.ruta_codigo = t3.ruta_codigo "
                "inner join planificacion_geocerca t4 on t3.id = t4.punto_control_id order by t1.numero_lpn")'''
            '''detalles = PlanificacionPuntoControl.objects.raw("select distinct t1.numero_lpn, t3.id, t3.nombre, t2.fecha_registro, t1.numero_carga, "
                "(select count(*) from planificacion_punto_control where ruta_codigo = t1.ruta_codigo ) numero_controles, "
                "t1.ruta_codigo, t1.local, t1.fecha_carga, "
                "t4.coordenadas_poligon polygon, t2.coordenadas_poligon point "
                "from planificacion_carga_bulto t1 "
                "left join planificacion_carga_punto_control t2 on t2.numero_lpn = t1.numero_lpn "
                "inner join planificacion_punto_control t3 on t1.ruta_codigo = t3.ruta_codigo "
                "inner join planificacion_geocerca t4 on t3.geocerca_id = t4.id "
                "where t3.ruta_codigo = %s "
                "or t1.numero_carga = %s "
                "or t1.numero_lpn = %s "
                "or t1.local = %s "
                "or (t1.fecha_carga BETWEEN convert(datetime2, %s, 103) AND convert(datetime2, %s, 103)) order by t1.numero_carga", [ruta,numero_carga,numero_bandeja,local,fecini,fecfin])'''
                
            detalles = PlanificacionCargaBulto.objects.raw("select t1.*, "
                "(select count(*) from planificacion_punto_control where ruta_codigo = t1.ruta_codigo ) numero_controles, "
                "t1.ruta_codigo, t1.local, t1.fecha_carga, "
                "t2.coordenadas_poligon point "
                "from planificacion_carga_bulto t1 "
                "left join planificacion_carga_punto_control t2 on t2.numero_lpn = t1.numero_lpn "
                "where t1.ruta_codigo = %s "
                "or t1.numero_carga = %s "
                "or t1.numero_lpn = %s "
                "or t1.local = %s "
                "or (t1.fecha_carga BETWEEN convert(datetime2, %s, 103) AND convert(datetime2, %s, 103)) order by t1.numero_carga", [ruta,numero_carga,numero_bandeja,local,fecini,fecfin])

            from collections import Counter
            #from shapely import wkt
            import math
            lpn1 = []
            '''m = 0
            while m < len(detalles):
                detalles[m].totales = 0
                detalles[m].distance = 0
                
                if detalles[m].polygon and detalles[m].point:
                    poly = wkt.loads('POLYGON('+str(detalles[m].polygon)+')')
                    pt = wkt.loads('POINT'+str(detalles[m].point)+'')
                    distance = poly.distance(pt)  # 0.0
                    detalles[m].distance = distance
                    if distance == 0.0:
                        lpn1.append(detalles[m].numero_lpn)
                        detalles[m].totales = int(detalles[m].totales + 1)
                        resum.append(detalles[m])
                m = m + 1

            resumen = dict(Counter(resum))
            #for key,value in resumen.items():
            #resum = resumen + a
            print(resumen)'''

            #------------ DETALLE-----------------#
            numCarga = []
            y = 0
            while y < len(detalles):
                detalles[y].display = 'table-row'
                if not detalles[y].numero_carga in numCarga:
                    numCarga.append(detalles[y].numero_carga)
                    data2.append(detalles[y])
                else:
                    detalles[y].display = 'none'
                    data2.append(detalles[y])
                y = y + 1


            lpn = []
            lpn2 = []
            n = 0
            while n < len(detalles):
                m = 0
                detalles[n].x = ''
                while m < len(dataPuntoControl):
                    if detalles[n].ruta_codigo == dataPuntoControl[m].ruta_codigo:
                        if dataPuntoControl[m].geocerca != None:
                            try:
                                poly = wkt.loads('POLYGON('+str(dataPuntoControl[m].geocerca.coordenadas_poligon)+')')
                                pt = wkt.loads('POINT'+str(detalles[n].point)+'')
                                distance = poly.distance(pt)  # 0.0
                                if distance == 0.0:
                                    print('yyyyyyy')
                                    print(detalles[n].numero_lpn)
                                    detalles[n].x = 'x'
                                    setattr(detalles[n], "orden", dataPuntoControl[m].orden)
                                    setattr(detalles[n], "control_"+str(dataPuntoControl[m].orden), 'x')
                                    data.append(detalles[n])
                                else:
                                    setattr(detalles[n], "orden", dataPuntoControl[m].orden)
                                    setattr(detalles[n], "control_"+str(dataPuntoControl[m].orden), '')
                                    data.append(detalles[n])
                            except Exception as e:
                                setattr(detalles[n], "orden", dataPuntoControl[m].orden)
                                setattr(detalles[n], "control_"+str(dataPuntoControl[m].orden), '')
                                data.append(detalles[n])
                        #else:
                        #    setattr(detalles[n], "control_"+str(m), '')
                        #    data.append(detalles[n])
                    m = m + 1
                n = n + 1

            '''lpn = []
            lpn2 = []
            n = 0
            while n < len(detalles):
                detalles[n].distance = 0
                detalles[n].x = 0
                if detalles[n].polygon and detalles[n].point:
                    poly = wkt.loads('POLYGON('+str(detalles[n].polygon)+')')
                    pt = wkt.loads('POINT'+str(detalles[n].point)+'')
                    distance = poly.distance(pt)  # 0.0
                    detalles[n].distance = distance
                    if distance == 0.0:
                        if not detalles[n].numero_lpn in lpn2:
                            lpn2.append(detalles[n].numero_lpn)
                            detalles[n].x = int(detalles[n].x + 1)
                            data.append(detalles[n])

                else:
                    if not detalles[n].numero_lpn in lpn2:
                        lpn2.append(detalles[n].numero_lpn)
                        data.append(detalles[n])
                n = n + 1'''


            
            '''numCarga = []
            y = 0
            while y < len(data):
                data[y].display = 'table-row'
                if not data[y].numero_carga in numCarga:
                    numCarga.append(data[y].numero_carga)
                    data2.append(data[y])
                else:
                    data[y].display = 'none'
                    data2.append(data[y])
                y = y + 1'''
            #--------------FIN DETALLE-----------#
            message = ''
            status = True
        except Exception as e:
            message = e
            status = False

    return render(request, 'consulta_registros.html', {"data_pc":data_pc,
        "countcargainicial":countcargainicial,
        "resumen":resumen,
        "resum":resum,
        "detalles":data2,
        "data":data,
        "message":message,
        "status":status,
        "numero_carga": numero_carga,
        "numero_bandeja": numero_bandeja,
        "ruta": ruta,
        "local": local,
        "fecini": fecini,
        "fecfin": fecfin})

@csrf_exempt
def export_consulta_registros(request):
    from django.db.models import Prefetch
    numero_carga = request.GET.get('numero_carga')
    numero_bandeja = request.GET.get('numero_bandeja')
    ruta = request.GET.get('ruta')
    local = request.GET.get('local')
    fecini = request.GET.get('fecini')
    fecfin = request.GET.get('fecfin')

    data_pc = {}
    countcargainicial = {}
    detalles = {}
    data = []
    resum = []

    if "GET" == request.method:
        from shapely import wkt
        
        data_pc = PlanificacionPuntoControl.objects.all().order_by('id')
        countcargainicial = PlanificacionCargaBulto.objects.all().count()
        '''resumen = PlanificacionPuntoControl.objects.raw("select distinct t1.numero_lpn,t3.id,t3.nombre, "
            "t3.id, t1.ruta_codigo, t2.coordenadas_poligon point, t4.coordenadas_poligon polygon "
            "from planificacion_carga_bulto t1 "
            "inner join planificacion_carga_punto_control t2 on t1.numero_lpn = t2.numero_lpn "
            "inner join planificacion_punto_control t3 on t1.ruta_codigo = t3.ruta_codigo "
            "inner join planificacion_geocerca t4 on t3.id = t4.punto_control_id order by t1.numero_lpn")'''
        detalles = PlanificacionPuntoControl.objects.raw("select distinct t1.numero_lpn, t3.id, t3.nombre, t2.fecha_registro, t1.numero_carga, "
            "(select count(*) from planificacion_punto_control where ruta_codigo = t1.ruta_codigo ) numero_controles, "
            "t1.ruta_codigo, t1.local, t1.fecha_carga, "
            "t4.coordenadas_poligon polygon, t2.coordenadas_poligon point "
            "from planificacion_carga_bulto t1 "
            "left join planificacion_carga_punto_control t2 on t2.numero_lpn = t1.numero_lpn "
            "inner join planificacion_punto_control t3 on t1.ruta_codigo = t3.ruta_codigo "
            "inner join planificacion_geocerca t4 on t3.geocerca_id = t4.id "
            "where t3.ruta_codigo = %s "
            "or t1.numero_carga = %s "
            "or t1.numero_lpn = %s "
            "or t1.local = %s "
            "or (t2.fecha_registro BETWEEN convert(datetime2, %s, 103) AND convert(datetime2, %s, 103)) order by t1.numero_lpn", [ruta,numero_carga,numero_bandeja,local,fecini,fecfin])

        from collections import Counter
        from shapely import wkt
        import math
        lpn1 = []
        m = 0
        while m < len(detalles):
            detalles[m].totales = 0
            detalles[m].distance = 0
            
            if detalles[m].polygon and detalles[m].point:
                poly = wkt.loads('POLYGON('+str(detalles[m].polygon)+')')
                pt = wkt.loads('POINT'+str(detalles[m].point)+'')
                distance = poly.distance(pt)  # 0.0
                detalles[m].distance = distance
                if distance == 0.0:
                    lpn1.append(detalles[m].numero_lpn)
                    detalles[m].totales = int(detalles[m].totales + 1)
                    resum.append(detalles[m])
            m = m + 1

        resumen = dict(Counter(resum))
        print('ppppppp')
        print(resumen)

        lpn = []
        n = 0
        while n < len(detalles):
            detalles[n].distance = 0
            detalles[n].x = 0
            if detalles[n].polygon and detalles[n].point:
                poly = wkt.loads('POLYGON('+str(detalles[n].polygon)+')')
                pt = wkt.loads('POINT'+str(detalles[n].point)+'')
                distance = poly.distance(pt)  # 0.0
                detalles[n].distance = distance
                if distance == 0.0:
                    if not detalles[n].numero_lpn in lpn:
                        lpn.append(detalles[n].numero_lpn)
                        detalles[n].x = int(detalles[n].x + 1)
                        data.append(detalles[n])
            else:
                if not detalles[n].numero_lpn in lpn:
                    lpn.append(detalles[n].numero_lpn)
                    data.append(detalles[n])
            n = n + 1

        # Export excel
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        #--------------RESUMEN---------------
        ws['A1'] = '-'
        ws['B1'] = 'Control 0'
        ws['A2'] = 'Total'
        ws['A3'] = 'Dentro de Geocerca'
        
        cr = 3
        for control in data_pc:
            ws.cell(row=1,column=cr).value = control.nombre
            cr = cr + 1

        ws.cell(row=2,column=2).value = countcargainicial
        
        for control in data_pc:
            x = 3
            for key,value in resumen.items():
                if key.id == control.id:
                    ws.cell(row=2,column=x).value = value
                x = x + 1

        for control in data_pc:
            x = 3
            for key,value in resumen.items():
                if key.id == control.id:
                    ws.cell(row=3,column=x).value = value
                x = x + 1
        

        #--------------DETALLE................
        ws['A5'] = 'Número de carga'
        ws['B5'] = 'Número de bandeja'
        ws['C5'] = '#Controles'
        ws['D5'] = 'Ruta'
        ws['E5'] = 'Local'
        ws['F5'] = 'Fecha de carga'

        c = 7
        for control in data_pc:
            ws.cell(row=5,column=c).value = control.nombre
            c = c + 1

        cont=6
        for row in data:
            ws.cell(row=cont,column=1).value = row.numero_carga
            ws.cell(row=cont,column=2).value = row.numero_lpn
            ws.cell(row=cont,column=3).value = row.numero_controles
            ws.cell(row=cont,column=4).value = row.ruta_codigo.codigo
            ws.cell(row=cont,column=5).value = row.local
            ws.cell(row=cont,column=6).value = row.fecha_carga
            column = 7
            for control in data_pc:
                if row.id == control.id:
                    if row.x == 1:
                        ws.cell(row=cont,column=column).value = 'x'
                    else:
                        ws.cell(row=cont,column=column).value = ''
                else:
                    ws.cell(row=cont,column=column).value = ''
                column = column + 1
            cont = cont + 1
        nombre_archivo ="ListadoFormasPago.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response 


@csrf_exempt
def registro_manual(request):
    if PuntoGeo.permisos(request.user.id).registro_manual:
        from django.db.models import Prefetch
        numero_carga = request.POST.get('numero_carga')
        numero_bandeja = request.POST.get('numero_bandeja')
        fecini = request.POST.get('fecini')
        fecfin = request.POST.get('fecfin')
        message = ''
        status = None
        data_pc = {}
        countcargainicial = {}
        resumen = {}
        detalles = {}
        detallado = {}
        data1 = []
        data = []
        if "POST" == request.method:
            #from shapely import wkt
            
            detalles = PlanificacionPuntoControl.objects.raw("select distinct t1.numero_lpn, t3.id, t3.nombre, t2.fecha_registro, t1.numero_carga, "
                "(select count(*) from planificacion_punto_control where ruta_codigo = t1.ruta_codigo ) numero_controles, "
                "t1.ruta_codigo, t1.local, t1.fecha_carga, "
                "t4.coordenadas_poligon polygon, t2.coordenadas_poligon point "
                "from planificacion_carga_bulto t1 "
                "left join planificacion_carga_punto_control t2 on t2.numero_lpn = t1.numero_lpn "
                "inner join planificacion_punto_control t3 on t1.ruta_codigo = t3.ruta_codigo "
                "inner join planificacion_geocerca t4 on t3.geocerca_id = t4.id "
                "where t1.numero_carga = %s "
                "or t1.numero_lpn = %s "
                "or (t1.fecha_carga BETWEEN convert(datetime2, %s, 103) AND convert(datetime2, %s, 103)) order by t1.numero_lpn", [numero_carga,numero_bandeja,fecini,fecfin])

            #from shapely import wkt
            lpn = []
            n = 0
            while n < len(detalles):
                detalles[n].distance = 0
                detalles[n].x = 0
                if detalles[n].point:
                    #poly = wkt.loads('POLYGON('+str(detalles[n].polygon)+')')
                    #pt = wkt.loads('POINT'+str(detalles[n].point)+'')
                    #distance = poly.distance(pt)  # 0.0
                    #detalles[n].distance = distance
                    #if distance == 0.0:
                    #    if not detalles[n].numero_lpn in lpn:
                    #        lpn.append(detalles[n].numero_lpn)
                    #        detalles[n].x = int(detalles[n].x + 1)
                    #        data.append(detalles[n])
                    print(detalles[n].point)
                else:
                    if not detalles[n].numero_lpn in lpn:
                        lpn.append(detalles[n].numero_lpn)
                        data1.append(detalles[n])
            
                n = n + 1

            usuario = request.user.username
            if request.user.is_staff:
                data = data1
            else:               
                i = 0
                while i < len(data1):
                    cb = PlanificacionCargaBulto.objects.get(numero_lpn = data1[i].numero_lpn)
                    if cb.transportista == usuario:
                        data.append(data1[i])
                    i = i + 1
                    
                    

            message = ''
            status = None

        return render(request, 'registro_manual.html', {
            "detalles":data,
            "message":message,
            "status":status})
    else:
        return render(request, 'permisos.html')

@csrf_exempt
def registro_manual_scaner(request):
    message = ''
    status = False
    data = {}
    numero_placa = request.GET.get('numero_placa')
    numero_lpn = request.GET.get('numero_bandeja')
    latitud = request.GET.get('latitud')
    longitud = request.GET.get('longitud')
    fecha_registro = request.GET.get('fecha_registro')

    import datetime
    import time
    now = datetime.datetime.now()
    fecha_registro = datetime.datetime.strptime(str(fecha_registro), '%Y-%m-%d %H:%M:%S')

    try:
        coordenadas = '{"lat": '+latitud+', "lng": '+longitud+'}'
        coordenadas_poligon = '('+latitud+' '+longitud+')'
        PlanificacionCargaPuntoControl.objects.create(numero_lpn=numero_lpn,
            numero_placa=numero_placa,fecha_registro=fecha_registro,
            latitud=latitud,longitud=longitud,
            fecha_envio=now,coordenadas=coordenadas,coordenadas_poligon=coordenadas_poligon)
        status = True
        message = 'Registro agregado correctamente.'
    except Exception as e:
        message = 'Error al insertar registro. ' + str(e)

    #return redirect('/registro_manual/')
    data = {'status': status, 'message': message}
    return JsonResponse(data)
    #return render(request, 'registro_manual.html', {"message":message,"status":status})

@csrf_exempt
def consulta_incidencias(request):
    if PuntoGeo.permisos(request.user.id).consulta_incidentes:
        message = ''
        status = None
        data = {}
        from django.db.models import Prefetch
        ruta = request.POST.get('ruta')
        transportista = request.POST.get('transportista')
        numero_placa = request.POST.get('numero_placa')
        fecini = request.POST.get('fecini')
        fecfin = request.POST.get('fecfin')

        try:
            if "POST" == request.method:
                '''data = PlanificacionIncidencia.objects.raw(
                    "select distinct t2.id, t1.user_id, t2.fecha_registro, t2.ruta, t2.descripcion from planificacion_carga_bulto t1 "
                    "inner join planificacion_incidencia t2 on t1.ruta_codigo = t2.ruta "
                    "inner join auth_user t3 on t1.user_id = t3.id "
                    "where t2.numero_placa = %s or t2.ruta = %s "
                    "or (t2.fecha_registro BETWEEN convert(datetime2, %s, 103) AND convert(datetime2, %s, 103)) "
                    "or t3.username like %s ", [numero_placa, ruta, fecini, fecfin, transportista])'''
                data = []
                data1 = PlanificacionIncidencia.objects.raw(
                    "select distinct t2.id, t3.username transportista, t2.fecha_registro, t2.ruta, t2.descripcion, t3.username from planificacion_ruta t1 "
                    "right join planificacion_incidencia t2 on t1.codigo = t2.ruta "
                    "left join auth_user t3 on t2.usuario_id = t3.id "
                    "where t2.numero_placa = %s or t2.ruta = %s "
                    "or (t2.fecha_registro BETWEEN convert(datetime2, %s, 103) AND convert(datetime2, %s, 103)) "
                    "or t3.username like %s ", [numero_placa, ruta, fecini, fecfin, transportista])
                
                usuario = request.user.username
                if request.user.is_staff:
                    data = data1
                else:
                    for d in data1: 
                       if d.transportista == usuario:
                           data.append(d)

        except Exception as e:
            print(e)
            message = ''
            status = None

        return render(request, 'consulta_incidencias.html', {"data":data,"message":message,"status":status})
    else:
        return render(request, 'permisos.html')